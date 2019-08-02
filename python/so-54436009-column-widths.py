from openpyxl import Workbook, load_workbook
#from openpyxl.utils import get_column_letter
#from openpyxl.worksheet.dimensions import SheetFormatProperties

### Simple script scribblings to help with the answer to SO post
### https://stackoverflow.com/questions/54436009

#wbP = load_workbook("c:\\temp\\payments-upload.xlsx")
#wsP = wbP.active
#wsP = wbP["Employees"]

wb1 = load_workbook("c:\\temp\\col-width-1.xlsx")
wb2 = load_workbook("c:\\temp\\col-width-2.xlsx")

ws1 = wb1.active
ws2 = wb2.active

print("WS1 sheet props: ", ws1.sheet_format.defaultColWidth)
print("WS2 sheet props: ", ws2.sheet_format.defaultColWidth)

#print(wb.sheetnames)
#print(ws.title)

# payments = wb.defined_names["Payments"]
# dests = payments.destinations

# for title, coord in dests:
#     ws = wb[title]
#     print(coord)

#col_b_w = wsP.column_dimensions['B'].width
#col_e_w = wsP.column_dimensions['E'].width
#print("Col B width: ", col_b_w)
#print("Col E width: ", col_e_w)

#wb = Workbook()
#ws1 = wb.active
#ws1.title = "SO_Test1"

#ws2 = wb.create_sheet("P4")

#ws1.column_dimensions['B'].width = 30
## My SO answer
#if (col_e_w is not None):
#    ws1.column_dimensions['E'].width = col_e_w

## A better answer, perhaps
#for k, cd in wsP.column_dimensions.items():
#    if cd is not None and cd.width != 0:
#        ws1.column_dimensions[k].width = cd.width

#col_e_w = wsP.column_dimensions['E'].width

#ws1.column_dimensions['E'].width = None

#ws1['A1'] = "Hello"
#ws1['B2'] = "World"
#ws2['c1'] = 2
#ws2['c2'] = 4
#ws2['c3'] = 8

#i = get_column_letter(1)
#ws1.column_dimensions[i].width = 40

#i = get_column_letter(2)
#ws1.column_dimensions[i].width = 20

#wb.save(filename = "c:\\temp\\so-test4.xlsx")
